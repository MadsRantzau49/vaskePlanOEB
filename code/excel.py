from openpyxl import load_workbook, Workbook

def generate_workbook_template(filename):
    wb = Workbook()
    wb.create_sheet(title="Sheet1")

    ws = wb.active

    ws.cell(row=1, column=1).value = "Navn"
    ws.cell(row=1, column=2).value = "Kampe Spillet"
    ws.cell(row=1, column=3).value = "Antal Gange vasket Tøj"
    ws.cell(row=1, column=4).value = "Retfærdig rækkefølge"


    wb.save(filename)
    return wb

def extract_excel(filename):
    try:
        wb = load_workbook(filename)
        ws = wb.active
        player_data = []
        for row in ws.iter_rows():
            if row[0].value is not None:
                player_data.append({
                    "name": row[0].value,  
                    "gameCounter": 0,  
                    "washed": row[2].value, 
                    "order": 0  
        })
        del player_data[0]
        print("All The Current Data Is Found Succesfully")
        return player_data
    
    except FileNotFoundError:
        wb = generate_workbook_template(filename)
        print("No Data Found, Another Template Is Generated")
        return []


def update_excel(data, filename):
    wb = load_workbook(filename)

    ws = wb.active

    # Write data starting from the second row
    for row_num, data in enumerate(data, start=2):
        ws.cell(row=row_num, column=1).value = data["name"]
        ws.cell(row=row_num, column=2).value = data["gameCounter"]
        ws.cell(row=row_num, column=3).value = data["washed"]     
        ws.cell(row=row_num, column=4).value = data["order"]          


    # Auto-size columns to fit content
    for col in range(1, 5):
        max_length = max(len(str(ws.cell(row=row_num, column=col).value)) for row_num in range(1, ws.max_row + 1))
        ws.column_dimensions[chr(64 + col)].width = max_length
    
    # Add filter on the columns
    ws.auto_filter.ref = ws.dimensions
    
    wb.save(filename)
